"""
Smoke tests for the financial-analysis Excel export.

Build the workbook from a seeded case and assert structural invariants —
all expected sheets present, headline values land in known cells, number
formats applied, filename sanitized.
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook


def _seed_case(tmp_path: Path, case_id: str = "case01") -> Path:
    """Lay down the on-disk artefacts the Excel exporter reads."""
    case_root = tmp_path / case_id
    for sub in ("raw", "parsed", "features", "agents", "chat"):
        (case_root / sub).mkdir(parents=True, exist_ok=True)

    (case_root / "manifest.json").write_text(json.dumps({
        "case_id": case_id,
        "company_name": "Acme Pte Ltd",
        "industry_hint": "logistics",
        "currency": "SGD",
        "uen": "201912345A",
    }), encoding="utf-8")

    fs_analytics = {
        "entity": {
            "name": "Acme Pte Ltd",
            "uen": "201912345A",
            "framework": "SFRS",
            "audited": True,
            "consolidated": False,
        },
        "perimeter": "company",
        "fys": ["FY2024", "FY2023"],
        "summary_ratios": {"interest_coverage": 1.2, "debt_equity": 3.5,
                           "ebitda_margin": 0.082, "pat_margin": 0.018},
        "by_fy": {
            "FY2023": {
                "fy": "FY2023", "currency": "SGD",
                "raw": {"revenue": 10_000_000, "ebitda": 1_200_000, "pat": 600_000,
                        "total_debt": 6_000_000, "total_equity": 2_000_000,
                        "cfo": 800_000, "total_assets": 9_000_000, "gross_profit": 3_000_000},
                "ratios": {
                    "gross_margin": 0.30, "ebitda_margin": 0.12, "pat_margin": 0.06,
                    "current_ratio": 1.4, "debt_equity": 3.0, "interest_coverage": 1.8,
                    "receivable_days": 45,
                },
            },
            "FY2024": {
                "fy": "FY2024", "currency": "SGD",
                "raw": {"revenue": 11_000_000, "ebitda": 900_000, "pat": 200_000,
                        "total_debt": 7_500_000, "total_equity": 2_100_000,
                        "cfo": 400_000, "total_assets": 10_500_000, "gross_profit": 3_100_000},
                "ratios": {
                    "gross_margin": 0.282, "ebitda_margin": 0.082, "pat_margin": 0.018,
                    "current_ratio": 1.44, "debt_equity": 3.5, "interest_coverage": 1.2,
                    "receivable_days": 52,
                },
            },
        },
        "trends": {
            "interest_coverage_yoy_pct": -0.333,
            "ebitda_margin_yoy_pct": -0.317,
            "revenue_growth_yoy": 0.10,
            "ebitda_growth_yoy": -0.25,
        },
        "review_flags": [],
    }
    (case_root / "features" / "fs_analytics.json").write_text(json.dumps(fs_analytics), encoding="utf-8")

    (case_root / "assessment_summary.json").write_text(json.dumps({
        "cards": [
            {
                "card_type": "FS",
                "summary_title": "FS Spread",
                "risks": [
                    {"id": "r1", "severity": "high", "message": "EBITDA margin collapsed YoY"},
                    {"id": "r2", "severity": "medium", "message": "Receivables ageing"},
                ],
                "strengths": [{"id": "s1", "message": "Positive CFO"}],
            },
        ],
        "cross_findings": [
            {"severity": "medium", "message": "Cash on SoFP differs from CFO", "source": "CROSS"},
        ],
    }), encoding="utf-8")

    # Merged statements
    merged = case_root / "parsed" / "financials" / "merged"
    merged.mkdir(parents=True, exist_ok=True)
    sofp = {
        "statement": "sofp",
        "statement_name": "Statement of Financial Position",
        "perimeter": "company",
        "fys": ["FY2023", "FY2024"],
        "rows": [
            {"row_type": "section_header", "label": "Assets", "section_path": ["Assets"], "indent_level": 0, "values": {}},
            {"row_type": "line", "label": "Cash and Cash Equivalents",
             "canonical_code": "bs_cash", "indent_level": 2,
             "values": {"FY2023": 1_566_827, "FY2024": 490_541}},
            {"row_type": "subtotal", "label": "Total Current Assets",
             "canonical_code": "bs_current_assets", "indent_level": 1,
             "values": {"FY2023": 2_000_000, "FY2024": 1_500_000}},
        ],
    }
    (merged / "sofp__company.json").write_text(json.dumps(sofp), encoding="utf-8")
    soci = {
        "statement": "soci", "statement_name": "P&L",
        "perimeter": "company", "fys": ["FY2023", "FY2024"],
        "rows": [
            {"row_type": "line", "label": "Revenue",
             "canonical_code": "pl_revenue",
             "values": {"FY2023": 10_000_000, "FY2024": 11_000_000}},
        ],
    }
    (merged / "soci__company.json").write_text(json.dumps(soci), encoding="utf-8")

    # Qualitative probes
    (case_root / "agents" / "qualitative.json").write_text(json.dumps({
        "probes": [
            {"id": "q1", "question": "Explain the EBITDA decline",
             "priority": "high", "theme": "profitability",
             "rationale": "EBITDA fell sharply YoY",
             "documents_requested": ["FY25 forecast"],
             "evidence_links": [],
             "management_meeting": True},
        ],
    }), encoding="utf-8")

    return case_root


def _store_at(tmp_path: Path):
    from core.cases.case_store import CaseStore
    return CaseStore(base_dir=str(tmp_path))


# ---- tests -----------------------------------------------------------------

def test_workbook_contains_all_expected_sheets(tmp_path):
    _seed_case(tmp_path)
    from core.report.excel_export import build_analytics_workbook
    wb = build_analytics_workbook("case01", store=_store_at(tmp_path))
    names = wb.sheetnames
    for expected in ("Summary", "Balance Sheet", "Income Statement",
                     "Ratios", "YoY Trends", "Findings", "Probes"):
        assert expected in names, f"missing sheet: {expected}"


def test_summary_sheet_carries_entity_and_latest_ratios(tmp_path):
    _seed_case(tmp_path)
    from core.report.excel_export import build_analytics_workbook
    wb = build_analytics_workbook("case01", store=_store_at(tmp_path))

    # Round-trip through openpyxl so we exercise the same load path the
    # user's Excel will use.
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb = load_workbook(buf)

    ws = wb["Summary"]
    assert "Acme Pte Ltd" in str(ws["A2"].value)

    # Sweep the Summary sheet for the latest-FY interest_coverage cell.
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Interest coverage":
                neighbour = ws.cell(row=cell.row, column=cell.column + 1)
                assert neighbour.value == 1.2
                assert "x" in (neighbour.number_format or "")
                found = True
    assert found, "Interest coverage row not found on Summary"


def test_balance_sheet_rows_and_amount_format(tmp_path):
    _seed_case(tmp_path)
    from core.report.excel_export import build_analytics_workbook
    wb = build_analytics_workbook("case01", store=_store_at(tmp_path))
    ws = wb["Balance Sheet"]

    # Header row at row 4: ["Line item", "Code", "FY2023", "FY2024"]
    assert ws.cell(row=4, column=1).value == "Line item"
    assert ws.cell(row=4, column=3).value == "FY2023"
    assert ws.cell(row=4, column=4).value == "FY2024"

    # Find the cash row and assert the FY2024 amount carries a numeric format.
    cash_row = None
    for r in range(5, 20):
        if "Cash and Cash Equivalents" in str(ws.cell(row=r, column=1).value or ""):
            cash_row = r
            break
    assert cash_row, "Cash row not found"
    fy2024_cell = ws.cell(row=cash_row, column=4)
    assert fy2024_cell.value == 490_541
    assert "#,##0" in fy2024_cell.number_format


def test_ratios_sheet_groups_and_pct_format(tmp_path):
    _seed_case(tmp_path)
    from core.report.excel_export import build_analytics_workbook
    wb = build_analytics_workbook("case01", store=_store_at(tmp_path))
    ws = wb["Ratios"]

    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Profitability" in labels
    assert "Liquidity" in labels
    assert "Leverage & Coverage" in labels
    assert "EBITDA margin" in labels

    # EBITDA margin row should carry a percentage number format.
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "EBITDA margin":
            cell = ws.cell(row=r, column=2)  # FY2023 column
            assert "0.0%" in cell.number_format or "%" in cell.number_format
            break
    else:
        pytest.fail("EBITDA margin row not found")


def test_findings_sheet_sorted_high_first(tmp_path):
    _seed_case(tmp_path)
    from core.report.excel_export import build_analytics_workbook
    wb = build_analytics_workbook("case01", store=_store_at(tmp_path))
    ws = wb["Findings"]
    # First data row (row 4 after header) should be the HIGH-severity risk.
    assert ws.cell(row=4, column=1).value == "HIGH"
    assert "EBITDA margin collapsed" in str(ws.cell(row=4, column=4).value)


def test_probes_sheet_has_question(tmp_path):
    _seed_case(tmp_path)
    from core.report.excel_export import build_analytics_workbook
    wb = build_analytics_workbook("case01", store=_store_at(tmp_path))
    ws = wb["Probes"]
    assert ws.cell(row=4, column=1).value == "HIGH"
    assert "EBITDA decline" in str(ws.cell(row=4, column=3).value)


def test_filename_is_sanitized(tmp_path):
    _seed_case(tmp_path)
    from core.report.excel_export import analytics_xlsx_filename
    name = analytics_xlsx_filename("case01", store=_store_at(tmp_path))
    assert name.endswith(".xlsx")
    assert "Acme" in name


def test_missing_analytics_raises(tmp_path):
    # Manifest only — no fs_analytics.json.
    case_root = tmp_path / "case02"
    case_root.mkdir()
    (case_root / "manifest.json").write_text(json.dumps({
        "case_id": "case02", "company_name": "Bare Corp",
    }), encoding="utf-8")

    from core.report.excel_export import build_analytics_workbook
    with pytest.raises(ValueError, match="No FS analytics"):
        build_analytics_workbook("case02", store=_store_at(tmp_path))
