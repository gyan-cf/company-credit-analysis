"""
Excel export of the full financial analysis dashboard.

Produces a multi-sheet `.xlsx` workbook that mirrors what the analyst sees on
the Financials page plus the assessment output. Sheets:

    1. Summary           — entity, FYs, headline ratios, verdicts
    2. Balance Sheet     — merged SoFP (company perimeter, all FYs)
    3. Income Statement  — merged SoCI
    4. Cash Flow         — merged SoCF
    5. Ratios            — grouped by Profitability / Liquidity / Leverage / WC
    6. YoY Trends        — period-over-period change for each ratio
    7. Findings          — consolidated risks across assessment cards
    8. Probes            — qualitative-agent management questions

Inputs read from disk (no recomputation):
    cases/<id>/manifest.json
    cases/<id>/features/fs_analytics.json
    cases/<id>/parsed/financials/merged/{sofp,soci,socf}__company.json
    cases/<id>/assessment_summary.json
    cases/<id>/agents/qualitative.json

Public entry: `build_analytics_workbook(case_id, case_store) -> openpyxl.Workbook`
and `analytics_xlsx_filename(case_id, case_store) -> str` for HTTP layer.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.cases.case_store import CaseStore


# ---- styling --------------------------------------------------------------

BRAND_NAVY = "1F3A5F"
BRAND_LILAC = "EDE6FF"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_NAVY)
SECTION_FILL = PatternFill("solid", fgColor=BRAND_LILAC)
ALT_ROW_FILL = PatternFill("solid", fgColor="F8FAFC")

WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BLACK_BOLD = Font(name="Calibri", size=11, bold=True, color="1F2937")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="4338CA")
NORMAL = Font(name="Calibri", size=11, color="1F2937")
MUTED = Font(name="Calibri", size=10, color="64748B", italic=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_AMOUNT = "#,##0;[Red](#,##0);-"
FMT_RATIO_X = "0.00\"x\";[Red](0.00\"x\");-"
FMT_PCT = "0.0%;[Red]-0.0%;-"
FMT_DAYS = "#,##0\" d\";[Red]-#,##0\" d\";-"
FMT_RATIO = "0.00;[Red]-0.00;-"


# ---- ratio + label mappings ----------------------------------------------

RATIO_LABELS: Dict[str, str] = {
    "gross_margin": "Gross margin",
    "ebitda_margin": "EBITDA margin",
    "ebit_margin": "EBIT margin",
    "pat_margin": "PAT margin",
    "return_on_assets": "Return on assets",
    "return_on_equity": "Return on equity",
    "current_ratio": "Current ratio",
    "quick_ratio": "Quick ratio",
    "cash_ratio": "Cash ratio",
    "debt_equity": "Debt / equity",
    "debt_ebitda": "Debt / EBITDA",
    "debt_assets": "Debt / assets",
    "equity_ratio": "Equity ratio",
    "interest_coverage": "Interest coverage",
    "cfo_to_debt": "CFO / debt",
    "fcf_to_debt": "FCF / debt",
    "receivable_days": "Receivable days",
    "payable_days": "Payable days",
    "inventory_days": "Inventory days",
    "asset_turnover": "Asset turnover",
    "cfo_pat": "CFO / PAT",
    "cash_to_assets": "Cash / total assets",
}

RATIO_GROUPS: List[Tuple[str, List[str]]] = [
    ("Profitability", [
        "gross_margin", "ebitda_margin", "ebit_margin", "pat_margin",
        "return_on_assets", "return_on_equity",
    ]),
    ("Liquidity", [
        "current_ratio", "quick_ratio", "cash_ratio",
    ]),
    ("Leverage & Coverage", [
        "debt_equity", "debt_ebitda", "debt_assets", "equity_ratio",
        "interest_coverage", "cfo_to_debt", "fcf_to_debt",
    ]),
    ("Working Capital & Efficiency", [
        "receivable_days", "payable_days", "inventory_days",
        "asset_turnover", "cfo_pat", "cash_to_assets",
    ]),
]

# How a ratio's number is formatted on the sheet.
PCT_RATIOS = {
    "gross_margin", "ebitda_margin", "ebit_margin", "pat_margin",
    "return_on_assets", "return_on_equity", "debt_assets", "equity_ratio",
    "cfo_to_debt", "fcf_to_debt", "cash_to_assets",
}
DAYS_RATIOS = {"receivable_days", "payable_days", "inventory_days"}
MULTIPLE_RATIOS = {
    "current_ratio", "quick_ratio", "cash_ratio",
    "debt_equity", "debt_ebitda", "interest_coverage",
    "asset_turnover", "cfo_pat",
}


# ---- helpers --------------------------------------------------------------

def _fy_year_key(fy: str) -> int:
    digits = "".join(ch for ch in str(fy) if ch.isdigit())
    return int(digits[-4:]) if len(digits) >= 4 else 0


def _sorted_fys(fys: Iterable[str]) -> List[str]:
    return sorted([fy for fy in fys if fy], key=_fy_year_key)


def _load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def _set_header_row(ws: Worksheet, row: int, cols: List[str], widths: Optional[List[int]] = None) -> None:
    for i, label in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = HEADER_FILL
        c.font = WHITE_BOLD
        c.alignment = CENTER
        c.border = THIN_BORDER
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _safe_sheet_title(raw: str, taken: List[str]) -> str:
    cleaned = re.sub(r"[\\/\*\?\[\]:]", " ", raw)[:31].strip() or "Sheet"
    base = cleaned
    i = 2
    while cleaned in taken:
        suffix = f" ({i})"
        cleaned = (base[: 31 - len(suffix)] + suffix)
        i += 1
    taken.append(cleaned)
    return cleaned


def _autosize(ws: Worksheet, max_col: int, *, max_width: int = 48, min_width: int = 10) -> None:
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        longest = 0
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > longest:
                longest = length
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


# ---- sheet writers --------------------------------------------------------

def _write_summary(
    wb: Workbook,
    manifest: Dict[str, Any],
    fs_data: Dict[str, Any],
    assessment: Dict[str, Any],
) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "4F46E5"

    entity = fs_data.get("entity") or {}
    name = entity.get("name") or manifest.get("company_name") or "Unknown"
    uen = entity.get("uen") or manifest.get("uen") or manifest.get("cin") or "—"
    fys = _sorted_fys(fs_data.get("fys") or [])
    perimeter = fs_data.get("perimeter") or "company"

    # Title block
    ws["A1"] = "Financial Analysis Dashboard"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=BRAND_NAVY)
    ws.merge_cells("A1:F1")
    ws["A2"] = name
    ws["A2"].font = Font(name="Calibri", size=14, bold=True, color="111827")
    ws.merge_cells("A2:F2")

    info = [
        ("UEN", uen),
        ("Industry", manifest.get("industry_hint") or "—"),
        ("Framework", entity.get("framework") or "—"),
        ("Perimeter", perimeter.title()),
        ("Audited", "Yes" if entity.get("audited") else ("No" if entity.get("audited") is False else "—")),
        ("Consolidated", "Yes" if entity.get("consolidated") else "No"),
        ("Currency", manifest.get("currency") or "SGD"),
        ("FYs analysed", ", ".join(fys) if fys else "—"),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (k, v) in enumerate(info, start=4):
        ws.cell(row=i, column=1, value=k).font = BLACK_BOLD
        ws.cell(row=i, column=2, value=v).alignment = LEFT
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    # Headline ratios block
    headline_start = 4 + len(info) + 2
    ws.cell(row=headline_start, column=1, value="Headline Ratios (latest FY)").font = SECTION_FONT
    ws.cell(row=headline_start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=headline_start, start_column=1, end_row=headline_start, end_column=6)

    by_fy = fs_data.get("by_fy") or {}
    latest_fy = fys[-1] if fys else None
    latest_ratios = (by_fy.get(latest_fy) or {}).get("ratios", {}) if latest_fy else {}

    header_row = headline_start + 1
    _set_header_row(ws, header_row, ["Metric", "Value", "Group"])
    headline_keys = [
        ("revenue_growth_yoy_pct", "Revenue growth YoY", "Profitability", FMT_PCT),
    ]
    for key in ("ebitda_margin", "pat_margin", "current_ratio", "interest_coverage",
                "debt_equity", "debt_ebitda", "cfo_to_debt"):
        group = next((g for g, ks in RATIO_GROUPS if key in ks), "—")
        headline_keys.append((key, RATIO_LABELS.get(key, key), group, _fmt_for(key)))

    row = header_row + 1
    trends = fs_data.get("trends") or {}
    for key, label, group, fmt in headline_keys:
        ws.cell(row=row, column=1, value=label).font = NORMAL
        val = trends.get(key) if key.endswith("_yoy_pct") else latest_ratios.get(key)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt
        c.alignment = RIGHT
        ws.cell(row=row, column=3, value=group).font = MUTED
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Verdict / cards block
    cards = assessment.get("cards") or []
    cross_findings = assessment.get("cross_findings") or []
    verdict_row = row + 1
    ws.cell(row=verdict_row, column=1, value="Assessment").font = SECTION_FONT
    ws.cell(row=verdict_row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=verdict_row, start_column=1, end_row=verdict_row, end_column=6)

    if cards:
        _set_header_row(ws, verdict_row + 1, ["Card", "Title", "Risks", "Strengths"])
        for i, card in enumerate(cards):
            r = verdict_row + 2 + i
            ws.cell(row=r, column=1, value=card.get("card_type") or "—").font = BLACK_BOLD
            ws.cell(row=r, column=2, value=card.get("summary_title") or "—").alignment = LEFT
            ws.cell(row=r, column=3, value=len(card.get("risks") or []))
            ws.cell(row=r, column=4, value=len(card.get("strengths") or []))
            for col in range(1, 5):
                ws.cell(row=r, column=col).border = THIN_BORDER

    if cross_findings:
        cf_row = verdict_row + 2 + len(cards) + 1
        ws.cell(row=cf_row, column=1, value=f"Cross-source findings: {len(cross_findings)}").font = MUTED
        ws.merge_cells(start_row=cf_row, start_column=1, end_row=cf_row, end_column=6)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A4"


def _fmt_for(ratio_key: str) -> str:
    if ratio_key in PCT_RATIOS:
        return FMT_PCT
    if ratio_key in DAYS_RATIOS:
        return FMT_DAYS
    if ratio_key in MULTIPLE_RATIOS:
        return FMT_RATIO_X
    return FMT_RATIO


def _write_statement(
    wb: Workbook,
    case_root: Path,
    statement: str,
    title: str,
) -> Optional[Worksheet]:
    """Write one merged statement block (sofp/soci/socf, company perimeter)."""
    merged = case_root / "parsed" / "financials" / "merged" / f"{statement}__company.json"
    block = _load_json(merged)
    if not block:
        return None

    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = "0E7490"
    rows = block.get("rows") or []
    fys = _sorted_fys(block.get("fys") or [])

    # Header
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=BRAND_NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(fys))

    sub = block.get("statement_name") or ""
    ws["A2"] = sub
    ws["A2"].font = MUTED
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(fys))

    header_cols = ["Line item", "Code", *fys]
    _set_header_row(ws, 4, header_cols)

    r = 5
    for row in rows:
        row_type = row.get("row_type") or "line"
        if row_type == "spacer":
            r += 1
            continue
        label = row.get("label") or ""
        indent = int(row.get("indent_level") or 0)
        prefix = "  " * indent
        code = row.get("canonical_code") or ""
        values = row.get("values") or {}

        c0 = ws.cell(row=r, column=1, value=f"{prefix}{label}")
        c1 = ws.cell(row=r, column=2, value=code)
        c0.alignment = LEFT
        c1.font = MUTED
        c1.alignment = LEFT

        is_header = row_type == "section_header"
        is_subtotal = row_type in ("subtotal", "total")
        if is_header:
            c0.font = BLACK_BOLD
            c0.fill = SECTION_FILL
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        elif is_subtotal:
            c0.font = BLACK_BOLD
        else:
            c0.font = NORMAL

        for i, fy in enumerate(fys, start=3):
            v = values.get(fy)
            cell = ws.cell(row=r, column=i, value=v)
            cell.number_format = FMT_AMOUNT
            cell.alignment = RIGHT
            if is_subtotal:
                cell.font = BLACK_BOLD
            if not is_header:
                cell.border = THIN_BORDER

        if not is_header:
            for col in (1, 2):
                ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22
    for i, _fy in enumerate(fys, start=3):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "C5"
    return ws


def _write_ratios(wb: Workbook, fs_data: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Ratios")
    ws.sheet_properties.tabColor = "059669"

    fys = _sorted_fys(fs_data.get("fys") or [])
    by_fy = fs_data.get("by_fy") or {}

    ws["A1"] = "Ratios"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=BRAND_NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(fys))

    header_cols = ["Metric", *fys, "Unit"]
    _set_header_row(ws, 3, header_cols)

    r = 4
    for group_title, keys in RATIO_GROUPS:
        # group header band
        ws.cell(row=r, column=1, value=group_title).font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(fys))
        r += 1
        for key in keys:
            label = RATIO_LABELS.get(key, key)
            unit = (
                "%" if key in PCT_RATIOS else
                "days" if key in DAYS_RATIOS else
                "x" if key in MULTIPLE_RATIOS else
                ""
            )
            ws.cell(row=r, column=1, value=label).font = NORMAL
            for i, fy in enumerate(fys, start=2):
                ratios = (by_fy.get(fy) or {}).get("ratios") or {}
                v = ratios.get(key)
                c = ws.cell(row=r, column=i, value=v)
                c.number_format = _fmt_for(key)
                c.alignment = RIGHT
                c.border = THIN_BORDER
            ws.cell(row=r, column=len(fys) + 2, value=unit).font = MUTED
            ws.cell(row=r, column=1).border = THIN_BORDER
            r += 1
        r += 1  # blank row between groups

    ws.column_dimensions["A"].width = 28
    for i, _fy in enumerate(fys, start=2):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions[get_column_letter(len(fys) + 2)].width = 8
    ws.freeze_panes = "B4"


def _write_trends(wb: Workbook, fs_data: Dict[str, Any]) -> None:
    trends = fs_data.get("trends") or {}
    if not trends:
        return
    ws = wb.create_sheet("YoY Trends")
    ws.sheet_properties.tabColor = "B45309"

    ws["A1"] = "Year-over-Year Trends"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=BRAND_NAVY)
    ws.merge_cells("A1:C1")

    _set_header_row(ws, 3, ["Metric", "YoY change", "Group"])
    r = 4
    for group_title, keys in RATIO_GROUPS:
        seen_in_group = False
        for key in keys:
            trend_key = f"{key}_yoy_pct"
            v = trends.get(trend_key)
            if v is None:
                continue
            if not seen_in_group:
                ws.cell(row=r, column=1, value=group_title).font = SECTION_FONT
                ws.cell(row=r, column=1).fill = SECTION_FILL
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
                r += 1
                seen_in_group = True
            ws.cell(row=r, column=1, value=RATIO_LABELS.get(key, key)).font = NORMAL
            c = ws.cell(row=r, column=2, value=v)
            c.number_format = FMT_PCT
            c.alignment = RIGHT
            ws.cell(row=r, column=3, value=group_title).font = MUTED
            for col in (1, 2, 3):
                ws.cell(row=r, column=col).border = THIN_BORDER
            r += 1

    # Raw growth (revenue / ebitda / pat) appended at the bottom
    raw_keys = [
        ("revenue_growth_yoy", "Revenue growth"),
        ("ebitda_growth_yoy", "EBITDA growth"),
        ("pat_growth_yoy", "PAT growth"),
    ]
    raw_present = [(k, lbl) for k, lbl in raw_keys if trends.get(k) is not None]
    if raw_present:
        r += 1
        ws.cell(row=r, column=1, value="Top-line growth").font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1
        for k, lbl in raw_present:
            ws.cell(row=r, column=1, value=lbl).font = NORMAL
            c = ws.cell(row=r, column=2, value=trends.get(k))
            c.number_format = FMT_PCT
            c.alignment = RIGHT
            ws.cell(row=r, column=3, value="Growth").font = MUTED
            for col in (1, 2, 3):
                ws.cell(row=r, column=col).border = THIN_BORDER
            r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A4"


def _write_findings(wb: Workbook, assessment: Dict[str, Any]) -> None:
    cards = assessment.get("cards") or []
    cross_findings = assessment.get("cross_findings") or []
    rows: List[Tuple[str, str, str, str]] = []  # (severity, source, kind, message)
    sev_order = {"high": 3, "medium": 2, "low": 1, "info": 0}
    for card in cards:
        card_type = card.get("card_type") or "—"
        for risk in (card.get("risks") or []):
            rows.append((
                (risk.get("severity") or "low").lower(),
                card_type,
                "Risk",
                risk.get("message") or "",
            ))
        for strength in (card.get("strengths") or []):
            rows.append(("info", card_type, "Strength", strength.get("message") or ""))
    for cf in cross_findings:
        rows.append((
            (cf.get("severity") or "low").lower(),
            cf.get("source") or "CROSS",
            "Cross-finding",
            cf.get("message") or "",
        ))
    if not rows:
        return

    rows.sort(key=lambda t: sev_order.get(t[0], 0), reverse=True)

    ws = wb.create_sheet("Findings")
    ws.sheet_properties.tabColor = "DC2626"
    ws["A1"] = "Findings"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=BRAND_NAVY)
    ws.merge_cells("A1:D1")

    _set_header_row(ws, 3, ["Severity", "Source", "Kind", "Message"])
    for i, (sev, src, kind, msg) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=sev.upper())
        ws.cell(row=i, column=2, value=src)
        ws.cell(row=i, column=3, value=kind)
        ws.cell(row=i, column=4, value=msg).alignment = LEFT_TOP
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = THIN_BORDER
        if sev == "high":
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="FECACA")
            ws.cell(row=i, column=1).font = Font(bold=True, color="991B1B")
        elif sev == "medium":
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="FED7AA")
            ws.cell(row=i, column=1).font = Font(bold=True, color="9A3412")
        elif sev == "low":
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="FEF3C7")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 85
    ws.freeze_panes = "A4"


def _write_probes(wb: Workbook, qualitative: Dict[str, Any]) -> None:
    probes = qualitative.get("probes") or []
    if not probes:
        return
    ws = wb.create_sheet("Probes")
    ws.sheet_properties.tabColor = "7C3AED"

    ws["A1"] = "Management Probe Questions"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=BRAND_NAVY)
    ws.merge_cells("A1:F1")

    _set_header_row(ws, 3, ["Priority", "Theme", "Question", "Rationale", "Documents Requested", "Mgmt Meeting"])
    prio_order = {"high": 3, "medium": 2, "low": 1}
    sorted_probes = sorted(
        probes,
        key=lambda p: prio_order.get((p.get("priority") or "").lower(), 0),
        reverse=True,
    )
    for i, p in enumerate(sorted_probes, start=4):
        prio = (p.get("priority") or "").lower()
        ws.cell(row=i, column=1, value=prio.upper() or "—")
        ws.cell(row=i, column=2, value=p.get("theme") or "—")
        ws.cell(row=i, column=3, value=p.get("question") or "").alignment = LEFT_TOP
        ws.cell(row=i, column=4, value=p.get("rationale") or "").alignment = LEFT_TOP
        docs = p.get("documents_requested") or []
        ws.cell(row=i, column=5, value="; ".join(str(d) for d in docs)).alignment = LEFT_TOP
        ws.cell(row=i, column=6, value="Yes" if p.get("management_meeting") else "No")
        for col in range(1, 7):
            ws.cell(row=i, column=col).border = THIN_BORDER
        if prio == "high":
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="FECACA")
            ws.cell(row=i, column=1).font = Font(bold=True, color="991B1B")
        elif prio == "medium":
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="FED7AA")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 14
    ws.freeze_panes = "A4"


# ---- public entry ---------------------------------------------------------

def build_analytics_workbook(case_id: str, store: Optional[CaseStore] = None) -> Workbook:
    """Build the full financial-analysis workbook for one case."""
    store = store or CaseStore()
    case_root = store._case_path(case_id)
    manifest = store.get_manifest(case_id)
    fs_data = store.load_features(case_id, "fs_analytics") or {}
    assessment = store.load_assessment_summary(case_id) or {}
    qualitative = store.load_agent_result(case_id, "qualitative") or {}

    if not fs_data:
        raise ValueError(
            f"No FS analytics on disk for case {case_id}. Run /analyze first."
        )

    wb = Workbook()
    # Drop the default 'Sheet' that openpyxl creates.
    wb.remove(wb.active)

    _write_summary(wb, manifest, fs_data, assessment)
    _write_statement(wb, case_root, "sofp", "Balance Sheet")
    _write_statement(wb, case_root, "soci", "Income Statement")
    _write_statement(wb, case_root, "socf", "Cash Flow")
    _write_ratios(wb, fs_data)
    _write_trends(wb, fs_data)
    _write_findings(wb, assessment)
    _write_probes(wb, qualitative)

    # Fallback if every optional sheet was empty (should not happen given the
    # fs_data guard above, but defend against an unexpected shape).
    if not wb.sheetnames:
        wb.create_sheet("Summary")

    return wb


def analytics_xlsx_filename(case_id: str, store: Optional[CaseStore] = None) -> str:
    """Suggested filename for HTTP Content-Disposition."""
    store = store or CaseStore()
    try:
        manifest = store.get_manifest(case_id)
    except FileNotFoundError:
        return f"{case_id}_analysis.xlsx"
    name = manifest.get("company_name") or case_id
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", name).strip("_")[:60] or case_id
    return f"{safe}_{datetime.now().strftime('%Y%m%d')}_analysis.xlsx"
